"""
Step 1: Fetch all highlights from Readwise API and write/update the XLSX.

Rules:
  - New highlights are appended as new rows.
  - Existing rows (matched by highlight_id) have their Readwise columns updated.
  - 'CommonBook' is recomputed on every row (new and existing) based on
    INCLUDE_CATEGORIES and EXCLUDE_TAGS in config.py — manage exclusions via
    Readwise tags rather than editing the spreadsheet manually.
  - 'updated' is set to Y whenever Readwise data OR CommonBook value changes.
  - The 'updated' column is NEVER touched on unchanged rows.
  - The sheet is formatted as a named Excel Table with filter dropdowns.
  - Safe to re-run at any time.

Usage:
    python step1_fetch.py
"""

import requests
import openpyxl
import logging
import signal
import sys
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pathlib import Path
from config import (
    READWISE_TOKEN, XLSX_PATH, SHEET_NAME,
    ALL_COLS, READWISE_COLS, CURATION_COLS,
    INCLUDE_CATEGORIES, EXCLUDE_TAGS,
    XLSX_SAVE_INTERVAL,
)

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

LOG_PATH = Path(__file__).parent / "step1_fetch.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Graceful shutdown
# ---------------------------------------------------------------------------

_shutdown = False

def _handle_signal(signum, frame):
    global _shutdown
    _shutdown = True
    log.warning("Shutdown requested — finishing current row then saving...")

TABLE_NAME  = "ReadwiseHighlights"
TABLE_STYLE = "TableStyleMedium9"

COL_WIDTHS = {
    "highlight_id":  14,
    "title":         35,
    "author":        22,
    "date_added":    22,
    "date_modified": 22,
    "tags":          25,
    "date":          12,
    "quote":         60,
    "note":          30,
    "location":      10,
    "location_type": 14,
    "source_url":    30,
    "readwise_url":  35,
    "category":      12,
    "book_id":       12,
    "CommonBook":    12,
    "updated":       10,
}

# ---------------------------------------------------------------------------
# Readwise API
# ---------------------------------------------------------------------------
BASE_URL = "https://readwise.io/api/v2/export/"
HEADERS  = {"Authorization": f"Token {READWISE_TOKEN}"}


def fetch_all_highlights():
    highlights, params = [], {}
    while True:
        resp = requests.get(BASE_URL, headers=HEADERS,
                            params={k: v for k, v in params.items() if v})
        resp.raise_for_status()
        data = resp.json()
        for book in data.get("results", []):
            meta = {
                "title":      book.get("title", ""),
                "author":     book.get("author", ""),
                "category":   book.get("category", ""),
                "source_url": book.get("source_url", "") or "",
                "book_id":    str(book.get("user_book_id", "")),
                "book_tags":  [t["name"] for t in book.get("book_tags", [])],
            }
            for h in book.get("highlights", []):
                if h.get("is_deleted"):
                    continue
                htags    = [t["name"] for t in h.get("tags", [])]
                all_tags = list(dict.fromkeys(htags + meta["book_tags"]))
                hi_at    = h.get("highlighted_at") or ""
                updated  = h.get("updated", "") or ""
                # Date fallback: highlighted_at → updated_at → today
                if hi_at:
                    date_str = hi_at[:10]
                elif updated:
                    date_str = updated[:10]
                else:
                    from datetime import date
                    date_str = date.today().isoformat()
                highlights.append({
                    "highlight_id":  str(h["id"]),
                    "title":         meta["title"],
                    "author":        meta["author"],
                    "date_added":    hi_at or updated,
                    "date_modified": updated,
                    "tags":          ", ".join(all_tags),
                    "date":          date_str,
                    "quote":         h.get("text", ""),
                    "note":          h.get("note", "") or "",
                    "location":      str(h.get("location", "")) if h.get("location") is not None else "",
                    "location_type": h.get("location_type", "") or "",
                    "source_url":    meta["source_url"],
                    "readwise_url":  h.get("readwise_url", "") or "",
                    "category":      meta["category"],
                    "book_id":       meta["book_id"],
                    # helper key for auto-inclusion logic — not written to sheet
                    "_tags_list":    all_tags,
                })
        cursor = data.get("nextPageCursor")
        if not cursor:
            break
        params["pageCursor"] = cursor
    return highlights


def auto_include(highlight: dict) -> str:
    """Return 'Y' if this highlight should be auto-marked for CommonBook, else ''."""
    in_category  = highlight["category"] in INCLUDE_CATEGORIES
    has_excl_tag = any(tag in highlight["_tags_list"] for tag in EXCLUDE_TAGS)
    return "Y" if in_category and not has_excl_tag else ""


# ---------------------------------------------------------------------------
# Table helpers
# ---------------------------------------------------------------------------

def remove_existing_table(ws):
    to_remove = [t for t in ws.tables.values() if t.name == TABLE_NAME]
    for t in to_remove:
        del ws.tables[t.name]


def apply_table(ws):
    n_cols = len(ALL_COLS)
    ref    = f"A1:{get_column_letter(n_cols)}{ws.max_row}"
    tbl    = Table(displayName=TABLE_NAME, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,   showColumnStripes=False,
    )
    ws.add_table(tbl)


def apply_column_widths(ws):
    for idx, col_name in enumerate(ALL_COLS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(col_name, 15)


# ---------------------------------------------------------------------------
# XLSX load / create
# ---------------------------------------------------------------------------

def load_or_create_workbook(path: str):
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if p.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        if ws.max_row < 1 or ws.cell(1, 1).value is None:
            ws.append(ALL_COLS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(ALL_COLS)
    return wb, ws


def build_id_row_map(ws) -> dict:
    id_col = ALL_COLS.index("highlight_id") + 1
    return {
        str(row[id_col - 1].value): row[0].row
        for row in ws.iter_rows(min_row=2)
        if row[id_col - 1].value is not None
    }


def write_row(ws, row_num: int, highlight: dict) -> bool:
    """Write Readwise columns only; return True if any value changed."""
    changed = False
    for col_idx, col_name in enumerate(ALL_COLS, start=1):
        if col_name in ("CommonBook", "updated"):
            continue                    # never overwrite curation columns
        if col_name not in READWISE_COLS:
            continue                    # skip any unrecognised key
        new_val = highlight.get(col_name, "")
        old_val = ws.cell(row=row_num, column=col_idx).value
        if (old_val or "") != (new_val or ""):
            ws.cell(row=row_num, column=col_idx, value=new_val)
            changed = True
    return changed


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    signal.signal(signal.SIGINT,  _handle_signal)
    signal.signal(signal.SIGTERM, _handle_signal)

    log.info("Fetching highlights from Readwise...")
    highlights = fetch_all_highlights()
    log.info(f"Retrieved {len(highlights)} highlights.")

    wb, ws = load_or_create_workbook(XLSX_PATH)
    id_map = build_id_row_map(ws)

    new_count = updated_count = skipped_count = 0
    rows_since_save = 0

    for h in highlights:
        hid = h["highlight_id"]
        if hid in id_map:
            changed = write_row(ws, id_map[hid], h)
            cb_col     = ALL_COLS.index("CommonBook") + 1
            upd_col    = ALL_COLS.index("updated") + 1
            new_cb     = auto_include(h)
            old_cb     = ws.cell(row=id_map[hid], column=cb_col).value or ""
            cb_changed = old_cb != new_cb
            if cb_changed:
                ws.cell(row=id_map[hid], column=cb_col, value=new_cb)
            if changed or cb_changed:
                ws.cell(row=id_map[hid], column=upd_col, value="Y")
                updated_count += 1
                log.info(f"Updated row: {h['highlight_id']} ({h['title'][:40]})")
            else:
                skipped_count += 1
        else:
            next_row = ws.max_row + 1
            write_row(ws, next_row, h)
            cb_col  = ALL_COLS.index("CommonBook") + 1
            upd_col = ALL_COLS.index("updated") + 1
            ws.cell(row=next_row, column=cb_col,  value=auto_include(h))
            ws.cell(row=next_row, column=upd_col, value="")
            id_map[hid] = next_row
            new_count += 1
            log.info(f"New row: {h['highlight_id']} ({h['title'][:40]})")

        rows_since_save += 1
        if XLSX_SAVE_INTERVAL > 0 and rows_since_save >= XLSX_SAVE_INTERVAL:
            wb.save(XLSX_PATH)
            log.info(f"Periodic save after {rows_since_save} rows.")
            rows_since_save = 0

        # Graceful shutdown check — finish this row, then stop
        if _shutdown:
            log.warning("Graceful shutdown: saving XLSX and exiting.")
            break

    remove_existing_table(ws)
    apply_table(ws)
    apply_column_widths(ws)

    wb.save(XLSX_PATH)
    log.info(f"Done. {new_count} new | {updated_count} updated | {skipped_count} unchanged.")
    log.info(f"Saved: {XLSX_PATH}")


if __name__ == "__main__":
    main()
