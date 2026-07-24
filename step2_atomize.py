"""
Step 2: Read the XLSX and write one Obsidian note per row where CommonBook == 'Y'.

Title format:  YYYY-MM-DD - RW -- Q{NN}
  NN is a zero-padded counter over all accepted quotes globally on that date,
  continuing from any existing files so re-runs never collide.

New note flow:
  1. Build YAML frontmatter + raw template body as a single string.
  2. Write to OBSIDIAN_STAGING_DIR (outside vault).
  3. Move into OBSIDIAN_QUOTES_DIR (inside vault) — Templater fires on the
     complete file with YAML already present.

Overwrite policy:
  - Note doesn't exist + CommonBook = Y  → always create.
  - Note exists + updated = Y + fields changed → auto-update; clear flag.
  - Note exists + updated = Y + no change     → skip; clear flag.
  - Note exists + updated blank               → skip silently.

Tracked fields (COMPARE_FIELDS): book title, author, tags, Quote, note,
  source, RW source, class, highlight id.

Shutdown: press Ctrl+C once to finish the current row, save XLSX, and exit.

Usage:
    python3 step2_atomize.py
"""

import sys
import re
import yaml
import shutil
import time
import signal
import logging
import openpyxl
from pathlib import Path
from collections import defaultdict
from config import (
    XLSX_PATH, SHEET_NAME,
    OBSIDIAN_QUOTES_DIR, OBSIDIAN_STAGING_DIR,
    QUOTES_TEMPLATE_PATH, STAGING_DELAY,
    INCLUDE_VALUE, COUNTER_PADDING, EXTRA_TAGS, OBS_NOTE_TYPE, OBS_VERSION,
    XLSX_SAVE_INTERVAL,
)
from columns import ALL_COLS, HEADER_TO_KEY, KEY_TO_HEADER

FILENAME_PATTERN = "{date} - RW -- Q{counter}"

COMPARE_FIELDS = {
    "book title", "author", "tags", "Quote",
    "note", KEY_TO_HEADER["source_url"], "RW source",
    "class", KEY_TO_HEADER["highlight_id"],
}

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

LOG_PATH = Path(__file__).parent / "step2_atomize.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Graceful shutdown
# ---------------------------------------------------------------------------

_shutdown = False


def _handle_signal(signum, frame):
    global _shutdown
    _shutdown = True
    log.warning("Shutdown requested — finishing current row then saving...")


# ---------------------------------------------------------------------------
# Template loading
# ---------------------------------------------------------------------------

def load_template() -> str:
    path = Path(QUOTES_TEMPLATE_PATH)
    if not path.exists():
        raise FileNotFoundError(f"Template not found at: {QUOTES_TEMPLATE_PATH}")
    return path.read_text(encoding="utf-8")


# ---------------------------------------------------------------------------
# Frontmatter + note rendering
# ---------------------------------------------------------------------------

def sanitize_tag(tag: str) -> str:
    """Coerce a Readwise tag into a valid Obsidian tag.

    Obsidian tags allow letters, digits, '_', '-' and '/' (nesting) only.
    Spaces and other characters are replaced with '-'; a leading '#' and
    stray leading/trailing separators are stripped.
    """
    t = tag.strip().lstrip("#")
    t = re.sub(r"\s+", "-", t)               # whitespace → hyphen
    t = re.sub(r"[^\w\-/]", "-", t)          # any other invalid char → hyphen
    t = re.sub(r"-{2,}", "-", t)             # collapse repeats
    return t.strip("-/")


def split_authors(raw: str) -> list:
    """Split a Readwise author string into individual authors.

    Handles the common separators (',', ';', '&', ' and ') and drops empties.
    """
    parts = re.split(r"\s*(?:,|;|&|\band\b)\s*", raw or "")
    return [p.strip() for p in parts if p.strip()]


def _is_empty(value) -> bool:
    """True for values that should be omitted from the frontmatter."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, list):
        return all(_is_empty(v) for v in value)
    return False


def build_frontmatter(row: dict, title: str) -> str:
    tags = [t.strip() for t in row.get("tags", "").split(",") if t.strip()]
    tags += EXTRA_TAGS
    tags = [t for t in (sanitize_tag(t) for t in tags) if t]
    tags = list(dict.fromkeys(tags))         # dedupe, preserve order

    book_title = row.get("title", "")
    data = {
        "title":             title,
        "book title":        f"[[{book_title}]]" if book_title else "",
        "author":            split_authors(row.get("author", "")),
        "tags":              tags,
        "obs note type":     OBS_NOTE_TYPE,
        "obs version":       OBS_VERSION,
        "Quote":             row.get("quote", ""),
        "note":              row.get("note", "") or "",
        KEY_TO_HEADER["source_url"]:    row.get("source_url", "") or "",
        "RW source":                    row.get("readwise_url", "") or "",
        "class":                        row.get("category", "") or "",
        KEY_TO_HEADER["highlight_id"]:  row.get("highlight_id", ""),
    }
    data = {k: v for k, v in data.items() if not _is_empty(v)}
    return yaml.dump(data, allow_unicode=True, sort_keys=False, default_flow_style=False)


def render_note(row: dict, title: str, template: str) -> str:
    fm = build_frontmatter(row, title)
    return f"---\n{fm}---\n{template}"


# ---------------------------------------------------------------------------
# Staging write (outside vault → inside vault)
# ---------------------------------------------------------------------------

def write_to_staging(title: str, content: str):
    staging_dir = Path(OBSIDIAN_STAGING_DIR)
    quotes_dir  = Path(OBSIDIAN_QUOTES_DIR)
    staging_dir.mkdir(parents=True, exist_ok=True)
    quotes_dir.mkdir(parents=True, exist_ok=True)
    staging_file = staging_dir / f"{title}.md"
    quotes_file  = quotes_dir  / f"{title}.md"
    staging_file.write_text(content, encoding="utf-8")
    shutil.move(str(staging_file), str(quotes_file))
    time.sleep(STAGING_DELAY)


# ---------------------------------------------------------------------------
# Field-scoped comparison
# ---------------------------------------------------------------------------

def extract_yaml_fields(content: str, fields: set) -> dict:
    match = re.match(r"^---\n(.*?)\n---", content, re.DOTALL)
    if not match:
        return {}
    try:
        fm = yaml.safe_load(match.group(1)) or {}
    except yaml.YAMLError:
        return {}
    return {k: fm.get(k) for k in fields}


# ---------------------------------------------------------------------------
# Counter helpers
# ---------------------------------------------------------------------------

def existing_counters_for_date(vault_dir: Path, date_str: str) -> set:
    pattern = re.compile(rf"^{re.escape(date_str)} - RW -- Q(\d+)\.md$")
    counters = set()
    for f in vault_dir.glob(f"{date_str} - RW -- Q*.md"):
        m = pattern.match(f.name)
        if m:
            counters.add(int(m.group(1)))
    return counters


def next_counter(used: set) -> int:
    n = 0
    while n in used:
        n += 1
    return n


# ---------------------------------------------------------------------------
# Vault file lookup
# ---------------------------------------------------------------------------

def find_existing_file(vault_dir: Path, date_str: str, highlight_id: str) -> Path | None:
    for f in vault_dir.glob(f"{date_str}*.md"):
        try:
            content = f.read_text(encoding="utf-8")
            hid_key = re.escape(KEY_TO_HEADER["highlight_id"])
            if re.search(rf"{hid_key}:\s*['\"]?{re.escape(highlight_id)}['\"]?", content):
                return f
        except Exception:
            pass
    return None


# ---------------------------------------------------------------------------
# XLSX loading
# ---------------------------------------------------------------------------

def load_accepted_rows(xlsx_path: str) -> tuple[object, object, list[dict], dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    headers   = [str(h) if h is not None else "" for h in next(rows_iter)]
    accepted, id_to_row = [], {}
    for row_num, row in enumerate(rows_iter, start=2):
        r = dict(zip(headers, (str(v) if v is not None else "" for v in row)))
        r = {HEADER_TO_KEY.get(k, k): v for k, v in r.items()}  # XLSX header → internal key
        if r.get("highlight_id"):
            id_to_row[r["highlight_id"]] = row_num
        if r.get("CommonBook", "").strip() == INCLUDE_VALUE:
            accepted.append(r)
    return wb, ws, accepted, id_to_row


# ---------------------------------------------------------------------------
# XLSX flag clearing
# ---------------------------------------------------------------------------

def clear_updated_flag(ws, row_num: int):
    upd_col = ALL_COLS.index("updated") + 1
    ws.cell(row=row_num, column=upd_col, value="")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    signal.signal(signal.SIGINT,  _handle_signal)
    signal.signal(signal.SIGTERM, _handle_signal)

    out_dir = Path(OBSIDIAN_QUOTES_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    template = load_template()
    wb, ws, rows, id_to_row = load_accepted_rows(XLSX_PATH)
    log.info(f"Found {len(rows)} rows marked '{INCLUDE_VALUE}' in CommonBook.")

    by_date = defaultdict(list)
    for r in rows:
        by_date[r.get("date", "nodate")].append(r)

    written = skipped = auto_updated = 0
    rows_since_save = 0

    for date_str, date_rows in sorted(by_date.items()):
        used_counters = existing_counters_for_date(out_dir, date_str)

        for row in date_rows:
            hid           = row["highlight_id"]
            is_updated    = row.get("updated", "").strip() == "Y"
            xlsx_row      = id_to_row.get(hid)
            existing_file = find_existing_file(out_dir, date_str, hid)

            if existing_file:
                if not is_updated:
                    skipped += 1
                else:
                    old_content = existing_file.read_text(encoding="utf-8")
                    new_content = render_note(row, existing_file.stem, template)
                    old_fields  = extract_yaml_fields(old_content, COMPARE_FIELDS)
                    new_fields  = extract_yaml_fields(new_content, COMPARE_FIELDS)

                    if old_fields != new_fields:
                        existing_file.write_text(new_content, encoding="utf-8")
                        auto_updated += 1
                        log.info(f"Auto-updated: {existing_file.name}")
                    else:
                        skipped += 1

                    if xlsx_row:
                        clear_updated_flag(ws, xlsx_row)

            else:
                n     = next_counter(used_counters)
                used_counters.add(n)
                pad   = str(n).zfill(COUNTER_PADDING)
                title = FILENAME_PATTERN.format(date=date_str, counter=pad)
                write_to_staging(title, render_note(row, title, template))
                written += 1
                log.info(f"Written: {title}.md")

            # Periodic XLSX save
            rows_since_save += 1
            if XLSX_SAVE_INTERVAL > 0 and rows_since_save >= XLSX_SAVE_INTERVAL:
                wb.save(XLSX_PATH)
                log.info(f"Periodic save after {rows_since_save} rows.")
                rows_since_save = 0

            # Graceful shutdown — finish this row then exit
            if _shutdown:
                log.warning("Graceful shutdown: saving XLSX and exiting.")
                break

        if _shutdown:
            break

    wb.save(XLSX_PATH)
    log.info("Final XLSX save complete.")
    log.info(f"Done. {written} new | {auto_updated} updated | {skipped} skipped")


if __name__ == "__main__":
    main()
